"""Bank of England Government Liability Curve (GLC) downloader.

The BoE publishes UK constant-maturity nominal sovereign-yield spot
curves as end-of-month spreadsheets bundled in a single ZIP archive on
their website. This module fetches that archive, iterates the three XLSX files inside (1970→2015,
2016 to 2024, 2025 to present), parses each one's "4. spot curve" sheet,
and extracts the column at the maturity requested per series.

HTTP-200-HTML trap guard: BoE serves error pages with HTTP 200 and
HTML content rather than 4xx/5xx, so we explicitly verify the downloaded
bytes begin with the literal ZIP signature ``PK\\x03\\x04`` before
parsing. An HTML error page would start with ``<`` and fail the check.

Both maturities (2.0 and 10.0 years) live in the same archive and the
same sheet; an in-process cache means downloading them back-to-back
incurs the network fetch only once.
"""

from __future__ import annotations

import io
import zipfile
from typing import Any

import openpyxl
import pandas as pd
import requests
from loguru import logger

from data.schemas.raw_fred import RAW_FRED_SCHEMA
from src.data.config import pipeline_config
from src.data.downloaders._common import (
    cache_raw_response,
    write_raw_csv,
    write_series_lineage,
)

# PK\x03\x04 is the first 4 bytes of every ZIP archive: used to defend against BoE's "HTTP 200 + HTML error page" trap.
ZIP_SIGNATURE = b"PK\x03\x04"
SPOT_SHEET_NAME = "4. spot curve"

# GLC "4. spot curve" layout (zero-indexed in openpyxl's iter_rows):
#   row 0: sheet title, row 1: blank, row 2: "Maturity",
#   row 3: "years:" plus the maturity headers (0.5, 1, 1.5, 2, … 25),
#   row 4+: a month-end date in col 0, spot yields at each maturity in cols 1+.

YEARS_HEADER_ROW_INDEX = 3
DATA_START_ROW_INDEX = 4

# In-process cache of archive bytes keyed by URL. Cleared explicitly in tests.
_archive_cache: dict[str, bytes] = {}


def download_boe_yc_series(series_name: str, *, save: bool = True) -> pd.DataFrame:
    """Download one maturity from the BoE GLC monthly archive.

    Args:
        series_name: Key under data_sources.boe_yc.series.
        save: If True, persist raw CSV, cache the archive, and write lineage.

    Returns:
        (date, value) DataFrame of month-end nominal spot yields in percent.
    """
    cfg = pipeline_config()["data_sources"]["boe_yc"]
    series_cfg = cfg["series"][series_name]

    maturity = float(series_cfg["maturity_years"])
    sheet_name = cfg.get("sheet_name", SPOT_SHEET_NAME)
    archive_url = cfg["archive_url"]

    logger.info(
        "BoE-YC {} (maturity={}y): fetching {}",
        series_name,
        maturity,
        archive_url,
    )

    archive_bytes = _fetch_archive(archive_url)
    if save:
        cache_raw_response("boe_yc", series_name, archive_bytes, "zip")

    df = _parse_archive(archive_bytes, maturity=maturity, sheet_name=sheet_name)
    RAW_FRED_SCHEMA.validate(df)

    if save:
        out = write_raw_csv(df, "boe_yc", series_name)
        write_series_lineage(
            source="boe_yc",
            series=series_name,
            code=f"GLC_{maturity}y_nominal_spot",
            raw_csv_path=out,
            transformations=[
                "http_get",
                "verify_zip_signature",
                "parse_glc_xlsx",
                "validate",
            ],
            parameters={
                "maturity_years": maturity,
                "sheet_name": sheet_name,
                "frequency": series_cfg["frequency"],
                "archive_url": archive_url,
            },
        )
    return df


def download_all_boe_yc() -> dict[str, pd.DataFrame]:
    """Download every series under data_sources.boe_yc."""
    cfg = pipeline_config()["data_sources"]["boe_yc"]["series"]
    return {name: download_boe_yc_series(name) for name in cfg}


def _fetch_archive(url: str) -> bytes:
    """Fetch the GLC zip and verify the ZIP signature.

    Uses an in-process cache so the second maturity downloaded in the same
    process reuses the bytes from the first call.

    Raises:
        RuntimeError: If the downloaded bytes do not start with the ZIP
            signature PK\\x03\\x04 i.e. BoE returned an HTTP-200 HTML
            error page rather than a real archive.
    """
    if url in _archive_cache:
        return _archive_cache[url]

    response = requests.get(
        url,
        timeout=120,
        headers={"User-Agent": "Mozilla/5.0 (uk-gdp-regime-forecasting)"},
    )
    response.raise_for_status()
    payload = response.content

    if not payload.startswith(ZIP_SIGNATURE):
        head = payload[:80]
        raise RuntimeError(
            f"BoE GLC archive at {url} did not return a ZIP "
            f"(first 4 bytes: {payload[:4]!r}). Likely an HTTP-200-HTML "
            f"error page. First 80 bytes: {head!r}"
        )
    _archive_cache[url] = payload
    return payload


def _parse_archive(
    archive_bytes: bytes,
    *,
    maturity: float,
    sheet_name: str,
) -> pd.DataFrame:
    """Parse every XLSX in the GLC archive into a (date, value) frame.

    Concatenates across all XLSX files in the zip, drops duplicates on
    date (keeping the last occurrence so refreshed periods stays), and
    returns sorted ascending by date.
    """
    frames: list[pd.DataFrame] = []
    with zipfile.ZipFile(io.BytesIO(archive_bytes)) as zf:
        for member in zf.namelist():
            if not member.lower().endswith(".xlsx"):
                continue
            with zf.open(member) as fh:
                xlsx_bytes = fh.read()
            frames.append(_parse_xlsx_bytes(xlsx_bytes, maturity=maturity, sheet_name=sheet_name))

    if not frames:
        raise RuntimeError("BoE GLC archive contained no XLSX files.")

    df = pd.concat(frames, ignore_index=True)
    # The archive's three XLSXs split history by period and don't usually overlap,
    # but if BoE re publishes a revised month in the "present" file, keep="last"
    # ensures the refreshed value is kept over the older snapshot from the historical file.
    df = df.drop_duplicates(subset=["date"], keep="last")
    df = df.sort_values("date").reset_index(drop=True)
    return df


def _parse_xlsx_bytes(
    xlsx_bytes: bytes,
    *,
    maturity: float,
    sheet_name: str,
) -> pd.DataFrame:
    """Parse a single GLC XLSX, returning (date, value) rows for maturity.

    The "4. spot curve" sheet layout:

    Row 3 (0-indexed): the years: header followed by maturities
      ``0.5, 1, 1.5, 2, 2.5, … 10, … 25``.
    Row 4+ (0-indexed): column 0 is the month-end date; subsequent
      columns are the spot yields at each maturity.

    The maturity is matched against the years header by exact float equality
    (tolerance ``1e-9``), so cell type (str vs float) does not matter.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(
            f"GLC XLSX has no sheet named {sheet_name!r}; " f"sheets found: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))

    if len(rows) <= YEARS_HEADER_ROW_INDEX:
        raise RuntimeError(
            f"GLC XLSX sheet {sheet_name!r} has only {len(rows)} rows; "
            f"expected at least {YEARS_HEADER_ROW_INDEX + 2}."
        )

    header = rows[YEARS_HEADER_ROW_INDEX]
    col_index = _find_maturity_column(header, maturity=maturity)

    dates: list[pd.Timestamp] = []
    values: list[float] = []
    for row in rows[DATA_START_ROW_INDEX:]:
        if col_index >= len(row):
            continue
        date_cell = row[0]
        value_cell = row[col_index]
        if date_cell is None or value_cell is None:
            continue
        try:
            value_f = float(value_cell)
        except (TypeError, ValueError):
            continue
        dates.append(pd.Timestamp(date_cell))
        values.append(value_f)

    return pd.DataFrame({"date": dates, "value": values})


def _find_maturity_column(header_row: tuple[Any, ...], *, maturity: float) -> int:
    """Return the column index whose header equals ``maturity``."""
    for i, cell in enumerate(header_row[1:], start=1):
        if cell is None:
            continue
        try:
            if abs(float(cell) - maturity) < 1e-9:
                return i
        except (TypeError, ValueError):
            continue
    raise RuntimeError(
        f"No column in GLC sheet header matches maturity {maturity} years. "
        f"Header values: {header_row[:30]}"
    )


__all__ = ["download_all_boe_yc", "download_boe_yc_series"]
