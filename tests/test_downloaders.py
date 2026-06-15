"""Unit and integration tests for the downloaders.

Unit tests use monkeypatch to redirect file output to tmp_path
and to fake HTTP responses; no network is touched. Integration tests are
marked @pytest.mark.integration and only run when invoked explicitly.
"""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import MagicMock

import pandas as pd
import pytest

import src.data.downloaders._common as common_mod
import src.data.downloaders.boe as boe_mod
import src.data.downloaders.boe_yc as boe_yc_mod
import src.data.downloaders.fred as fred_mod
import src.data.downloaders.ons as ons_mod
from src.data.downloaders._common import PENDING_SENTINEL

# Sentinel rejection


def _fake_ons_cfg(cdid: str) -> dict:
    return {
        "data_sources": {
            "ons": {
                "api_base": "https://www.ons.gov.uk",
                "series": {
                    "test_series": {
                        "cdid": cdid,
                        "path": "test/path",
                        "dataset": "testds",
                        "frequency": "quarterly",
                    },
                },
            },
        },
    }


def _fake_boe_cfg(code: str, url: str) -> dict:
    return {
        "data_sources": {
            "boe": {
                "iadb_base": "https://example.test/",
                "series": {
                    "test_series": {
                        "code": code,
                        "url": url,
                        "frequency": "monthly",
                    },
                },
            },
        },
    }


def _fake_fred_cfg(code: str) -> dict:
    return {
        "data_sources": {
            "fred": {
                "api_base": "https://api.stlouisfed.org/fred/series/observations",
                "api_key_env": "FRED_API_KEY",
                "rate_limit_sleep_seconds": 0.0,
                "series": {
                    "test_series": {
                        "code": code,
                        "frequency": "monthly",
                    },
                },
            },
        },
    }


def test_ons_refuses_pending_sentinel(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(ons_mod, "pipeline_config",
                        lambda: _fake_ons_cfg(PENDING_SENTINEL))
    with pytest.raises(RuntimeError, match="unresolved"):
        ons_mod.download_ons_series("test_series", save=False)


def test_boe_refuses_pending_url(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(boe_mod, "pipeline_config",
                        lambda: _fake_boe_cfg("CODE", PENDING_SENTINEL))
    with pytest.raises(RuntimeError, match="unresolved"):
        boe_mod.download_boe_series("test_series", save=False)


def test_fred_refuses_pending_sentinel(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(fred_mod, "pipeline_config",
                        lambda: _fake_fred_cfg(PENDING_SENTINEL))
    monkeypatch.setenv("FRED_API_KEY", "test_key")
    with pytest.raises(RuntimeError, match="unresolved"):
        fred_mod.download_fred_series("test_series", save=False)


def test_fred_refuses_when_api_key_missing(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(fred_mod, "pipeline_config",
                        lambda: _fake_fred_cfg("ANY"))
    monkeypatch.delenv("FRED_API_KEY", raising=False)
    monkeypatch.setattr(fred_mod, "load_dotenv", lambda *_a, **_kw: None)
    with pytest.raises(RuntimeError, match="FRED_API_KEY"):
        fred_mod.download_fred_series("test_series", save=False)


# Parsers


ONS_QUARTERLY_PAYLOAD = {
    "description": {"title": "Test"},
    "quarters": [
        {"date": "2020 Q1", "value": "-2.0"},
        {"date": "2020 Q2", "value": "-19.4"},
        {"date": "2020 Q3", "value": "17.0"},
    ],
    "months": [],
    "years": [],
}

ONS_MONTHLY_PAYLOAD = {
    "quarters": [],
    "months": [
        {"date": "2020 JAN", "value": "100.0"},
        {"date": "2020 FEB", "value": "101.0"},
    ],
    "years": [],
}

FRED_PAYLOAD = {
    "observations": [
        {"date": "2020-01-01", "value": "50.0"},
        {"date": "2020-02-01", "value": "55.0"},
        {"date": "2020-03-01", "value": "."},
    ],
}


def test_ons_parser_quarterly() -> None:
    df = ons_mod._parse_ons_response(ONS_QUARTERLY_PAYLOAD, "quarterly")
    assert list(df["date"]) == [
        pd.Timestamp("2020-03-31"),
        pd.Timestamp("2020-06-30"),
        pd.Timestamp("2020-09-30"),
    ]
    assert list(df["value"]) == [-2.0, -19.4, 17.0]


def test_ons_parser_monthly() -> None:
    df = ons_mod._parse_ons_response(ONS_MONTHLY_PAYLOAD, "monthly")
    assert df["date"].iloc[0] == pd.Timestamp("2020-01-01")
    assert df["value"].iloc[1] == 101.0


def test_ons_parser_rejects_unknown_frequency() -> None:
    with pytest.raises(ValueError, match="frequency"):
        ons_mod._parse_ons_response(ONS_QUARTERLY_PAYLOAD, "weekly")


def test_ons_parser_rejects_empty_array() -> None:
    with pytest.raises(RuntimeError, match="quarters"):
        ons_mod._parse_ons_response({"quarters": []}, "quarterly")


def test_fred_parser_drops_missing_value_dot() -> None:
    df = fred_mod._parse_fred_response(FRED_PAYLOAD)
    assert len(df) == 2
    assert df["value"].iloc[0] == 50.0


def test_fred_parser_rejects_empty_observations() -> None:
    with pytest.raises(RuntimeError, match="observations"):
        fred_mod._parse_fred_response({"observations": []})


def test_boe_parser_normalises_iadb_csv() -> None:
    text = "DATE,IUDBEDR\n01 Jan 2020,0.75\n01 Feb 2020,0.50\n"
    df = boe_mod._parse_iadb_csv(text)
    assert list(df.columns) == ["date", "value"]
    assert df["date"].iloc[0] == pd.Timestamp("2020-01-01")
    assert df["value"].iloc[1] == 0.5


# End-to-end with mocks: artefacts (CSV, lineage, cache) are written


def _redirect_outputs(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    monkeypatch.setattr(common_mod, "repo_root", lambda: tmp_path)


def test_ons_end_to_end_writes_artefacts(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _redirect_outputs(monkeypatch, tmp_path)
    monkeypatch.setattr(
        ons_mod,
        "pipeline_config",
        lambda: {
            "data_sources": {
                "ons": {
                    "api_base": "https://www.ons.gov.uk",
                    "series": {
                        "gdp_growth": {
                            "cdid": "IHYQ",
                            "path": "economy/grossdomesticproductgdp",
                            "dataset": "qna",
                            "frequency": "quarterly",
                        },
                    },
                },
            },
        },
    )

    fake_response = MagicMock()
    fake_response.content = json.dumps(ONS_QUARTERLY_PAYLOAD).encode("utf-8")
    fake_response.json.return_value = ONS_QUARTERLY_PAYLOAD
    fake_response.raise_for_status = MagicMock()
    monkeypatch.setattr(ons_mod.requests, "get",
                        lambda *_a, **_kw: fake_response)

    df = ons_mod.download_ons_series("gdp_growth", save=True)
    assert len(df) == 3

    csv_path = tmp_path / "data" / "raw" / "ons" / "gdp_growth.csv"
    assert csv_path.exists()

    lineage_path = tmp_path / "data" / "lineage" / "ons__gdp_growth.lineage.json"
    assert lineage_path.exists()
    record = json.loads(lineage_path.read_text(encoding="utf-8"))
    assert record["source"] == "ons:IHYQ"
    assert "validate" in record["transformations"]

    cache_dir = tmp_path / "data" / "raw" / "_api_responses" / "ons"
    cached = list(cache_dir.glob("gdp_growth__*.json"))
    assert len(cached) == 1


def test_fred_end_to_end_writes_artefacts(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _redirect_outputs(monkeypatch, tmp_path)
    monkeypatch.setenv("FRED_API_KEY", "test_key")
    monkeypatch.setattr(fred_mod, "load_dotenv", lambda *_a, **_kw: None)
    monkeypatch.setattr(fred_mod, "pipeline_config",
                        lambda: _fake_fred_cfg("DCOILBRENTEU"))

    fake_response = MagicMock()
    fake_response.content = json.dumps(FRED_PAYLOAD).encode("utf-8")
    fake_response.json.return_value = FRED_PAYLOAD
    fake_response.raise_for_status = MagicMock()
    monkeypatch.setattr(fred_mod.requests, "get",
                        lambda *_a, **_kw: fake_response)

    df = fred_mod.download_fred_series("test_series", save=True)
    assert len(df) == 2

    csv_path = tmp_path / "data" / "raw" / "fred" / "test_series.csv"
    assert csv_path.exists()
    lineage_path = tmp_path / "data" / "lineage" / "fred__test_series.lineage.json"
    assert lineage_path.exists()
    record = json.loads(lineage_path.read_text(encoding="utf-8"))
    assert "api_key" not in record["parameters"]  # never log secrets


# Cache pruning


def test_cache_raw_response_keeps_only_last_five(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    _redirect_outputs(monkeypatch, tmp_path)
    for i in range(7):
        common_mod.cache_raw_response(
            "ons", "x", f"payload{i}".encode(), "json")

    cache_dir = tmp_path / "data" / "raw" / "_api_responses" / "ons"
    files = sorted(cache_dir.glob("x__*.json"))
    assert len(files) == 5


# Integration: real network (manual only)


@pytest.mark.integration
def test_ons_gdp_growth_real_download() -> None:
    df = ons_mod.download_ons_series("gdp_growth", save=False)
    assert len(df) > 80


@pytest.mark.integration
def test_fred_brent_oil_real_download() -> None:
    df = fred_mod.download_fred_series("brent_oil", save=False)
    assert len(df) > 100


# BoE Yield Curve (boe_yc): zip archive of GLC monthly spot-curve XLSX files


def _fake_boe_yc_cfg() -> dict:
    # Synthetic boe_yc block exercising both maturities; archive URL is dummy
    # because tests mock the HTTP fetch.
    return {
        "data_sources": {
            "boe_yc": {
                "archive_url": "https://example.test/glc.zip",
                "sheet_name": "4. spot curve",
                "series": {
                    "gilt_2y_yield": {
                        "maturity_years": 2.0,
                        "frequency": "monthly",
                        "aggregation": "end_of_period",
                    },
                    "gilt_10y_yield": {
                        "maturity_years": 10.0,
                        "frequency": "monthly",
                        "aggregation": "end_of_period",
                    },
                },
            },
        },
    }


def _make_synthetic_glc_xlsx(
    years_cols: list[float],
    dates: list,
    values_by_maturity: dict,
) -> bytes:
    # Build a 1-sheet XLSX mirroring the GLC "4. spot curve" layout:
    #   row 0: title, row 1: blank, row 2: "Maturity",
    #   row 3: "years:" plus maturity headers,
    #   row 4+: month-end date + spot yields at each maturity.
    import io as _io

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "4. spot curve"
    ws.cell(row=1, column=2, value="UK nominal spot curve")
    ws.cell(row=3, column=1, value="Maturity")
    ws.cell(row=4, column=1, value="years:")
    for i, m in enumerate(years_cols, start=2):
        ws.cell(row=4, column=i, value=m)
    for r_idx, date in enumerate(dates, start=5):
        ws.cell(row=r_idx, column=1, value=date)
        for c_idx, m in enumerate(years_cols, start=2):
            ws.cell(row=r_idx, column=c_idx,
                    value=values_by_maturity[m][r_idx - 5])

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_synthetic_archive(xlsx_files: list) -> bytes:
    # Pack one or more (filename, bytes) pairs into a ZIP archive in memory.
    import io as _io
    import zipfile

    buf = _io.BytesIO()
    with zipfile.ZipFile(buf, mode="w") as zf:
        for name, content in xlsx_files:
            zf.writestr(name, content)
    return buf.getvalue()


def test_boe_yc_rejects_html_response_with_zip_signature_guard(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # BoE serves error pages with HTTP 200 plus HTML body; the downloader must
    # refuse anything whose first four bytes are not the ZIP signature.
    boe_yc_mod._archive_cache.clear()
    monkeypatch.setattr(boe_yc_mod, "pipeline_config",
                        lambda: _fake_boe_yc_cfg())

    fake = MagicMock()
    fake.content = b"<!DOCTYPE html><html><body>Service unavailable</body></html>"
    fake.raise_for_status = MagicMock()
    monkeypatch.setattr(boe_yc_mod.requests, "get", lambda *_a, **_kw: fake)

    with pytest.raises(RuntimeError, match="ZIP|HTML"):
        boe_yc_mod.download_boe_yc_series("gilt_2y_yield", save=False)


def test_boe_yc_parses_synthetic_archive_for_both_maturities(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    boe_yc_mod._archive_cache.clear()
    monkeypatch.setattr(boe_yc_mod, "pipeline_config",
                        lambda: _fake_boe_yc_cfg())

    years_cols = [0.5, 1.0, 1.5, 2.0, 2.5, 5.0, 10.0]
    dates = [pd.Timestamp("2020-01-31"), pd.Timestamp("2020-02-29")]
    values_by_maturity = {
        0.5: [0.10, 0.11],
        1.0: [0.20, 0.21],
        1.5: [0.30, 0.31],
        2.0: [0.40, 0.50],  # 2-year column
        2.5: [0.55, 0.56],
        5.0: [1.20, 1.22],
        10.0: [1.50, 1.70],  # 10-year column
    }
    xlsx = _make_synthetic_glc_xlsx(years_cols, dates, values_by_maturity)
    archive = _make_synthetic_archive([("synth.xlsx", xlsx)])

    fake = MagicMock()
    fake.content = archive
    fake.raise_for_status = MagicMock()
    monkeypatch.setattr(boe_yc_mod.requests, "get", lambda *_a, **_kw: fake)

    df2 = boe_yc_mod.download_boe_yc_series("gilt_2y_yield", save=False)
    assert list(df2["date"]) == dates
    assert list(df2["value"]) == [0.40, 0.50]

    df10 = boe_yc_mod.download_boe_yc_series("gilt_10y_yield", save=False)
    assert list(df10["value"]) == [1.50, 1.70]


def test_boe_yc_concatenates_across_multiple_xlsx_files(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # Mirrors the real archive's three-file split; concatenated rows should
    # cover both periods without duplication.
    boe_yc_mod._archive_cache.clear()
    monkeypatch.setattr(boe_yc_mod, "pipeline_config",
                        lambda: _fake_boe_yc_cfg())

    years_cols = [2.0, 10.0]
    xlsx_old = _make_synthetic_glc_xlsx(
        years_cols,
        [pd.Timestamp("2019-12-31")],
        {2.0: [0.30], 10.0: [1.20]},
    )
    xlsx_new = _make_synthetic_glc_xlsx(
        years_cols,
        [pd.Timestamp("2020-01-31")],
        {2.0: [0.40], 10.0: [1.50]},
    )
    archive = _make_synthetic_archive(
        [("old.xlsx", xlsx_old), ("new.xlsx", xlsx_new)])

    fake = MagicMock()
    fake.content = archive
    fake.raise_for_status = MagicMock()
    monkeypatch.setattr(boe_yc_mod.requests, "get", lambda *_a, **_kw: fake)

    df = boe_yc_mod.download_boe_yc_series("gilt_2y_yield", save=False)
    assert len(df) == 2
    assert list(df["date"]) == [pd.Timestamp(
        "2019-12-31"), pd.Timestamp("2020-01-31")]
    assert list(df["value"]) == [0.30, 0.40]


def test_boe_yc_raises_when_maturity_missing_from_header(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    boe_yc_mod._archive_cache.clear()
    # Build an archive whose XLSX header has no 10-year column; the 10y
    # downloader should raise rather than silently return nothing.
    monkeypatch.setattr(boe_yc_mod, "pipeline_config",
                        lambda: _fake_boe_yc_cfg())
    years_cols = [0.5, 1.0, 2.0]  # no 10.0
    xlsx = _make_synthetic_glc_xlsx(
        years_cols,
        [pd.Timestamp("2020-01-31")],
        {0.5: [0.1], 1.0: [0.2], 2.0: [0.4]},
    )
    archive = _make_synthetic_archive([("synth.xlsx", xlsx)])

    fake = MagicMock()
    fake.content = archive
    fake.raise_for_status = MagicMock()
    monkeypatch.setattr(boe_yc_mod.requests, "get", lambda *_a, **_kw: fake)

    with pytest.raises(RuntimeError, match="No column.*maturity 10"):
        boe_yc_mod.download_boe_yc_series("gilt_10y_yield", save=False)


def test_boe_yc_e2e_writes_artefacts(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    _redirect_outputs(monkeypatch, tmp_path)
    boe_yc_mod._archive_cache.clear()
    monkeypatch.setattr(boe_yc_mod, "pipeline_config",
                        lambda: _fake_boe_yc_cfg())

    xlsx = _make_synthetic_glc_xlsx(
        [2.0, 10.0],
        [pd.Timestamp("2020-01-31")],
        {2.0: [0.40], 10.0: [1.50]},
    )
    archive = _make_synthetic_archive([("synth.xlsx", xlsx)])
    fake = MagicMock()
    fake.content = archive
    fake.raise_for_status = MagicMock()
    monkeypatch.setattr(boe_yc_mod.requests, "get", lambda *_a, **_kw: fake)

    df = boe_yc_mod.download_boe_yc_series("gilt_2y_yield", save=True)
    assert len(df) > 0

    # Raw CSV, lineage, and cached archive zip should all be written
    csv_path = tmp_path / "data" / "raw" / "boe_yc" / "gilt_2y_yield.csv"
    assert csv_path.exists()

    lineage_path = tmp_path / "data" / "lineage" / \
        "boe_yc__gilt_2y_yield.lineage.json"
    assert lineage_path.exists()
    record = json.loads(lineage_path.read_text(encoding="utf-8"))
    assert record["source"] == "boe_yc:GLC_2.0y_nominal_spot"
    assert "verify_zip_signature" in record["transformations"]

    cache_dir = tmp_path / "data" / "raw" / "_api_responses" / "boe_yc"
    cached = list(cache_dir.glob("gilt_2y_yield__*.zip"))
    assert len(cached) == 1


@pytest.mark.integration
def test_boe_yc_gilt_2y_real_download() -> None:
    boe_yc_mod._archive_cache.clear()
    df = boe_yc_mod.download_boe_yc_series("gilt_2y_yield", save=False)
    # Archive covers 1970 onwards; we expect at least 300 monthly rows.
    assert len(df) > 300


@pytest.mark.integration
def test_boe_yc_gilt_10y_real_download() -> None:
    boe_yc_mod._archive_cache.clear()
    df = boe_yc_mod.download_boe_yc_series("gilt_10y_yield", save=False)
    assert len(df) > 300
